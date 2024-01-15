import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class XlsTool:
    def __init__(self, file_path):
        self.file_path = file_path
        if os.path.exists(file_path):
            self.workbook = load_workbook(file_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active

    def write_header(self, headers):
        for col_index, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=1, column=col_index, value=header)
            cell.font = Font(bold=True)

    def insert_data(self, data):
        if isinstance(data, list):
            for row_data in data:
                self.sheet.append(row_data)
        else:
            self.sheet.append(data)

    def save(self):
        self.workbook.save(self.file_path)
        print(f"Excel file '{self.file_path}' has been created.")


#
if __name__ == '__main__':
    xls_tool = XlsTool('output.xlsx')

    # 写表头
    header_data = ['Name', 'Age', 'City']
    xls_tool.write_header(header_data)

    # 插入数据
    data_to_insert = [['John', 25, 'New York'],
                      ['Jane', 30, 'Los Angeles'],
                      ['Bob', 22, 'Chicago']]
    xls_tool.insert_data(data_to_insert)

    # 保存文件
    xls_tool.save()
